#!/usr/bin/env python3
"""Turn the two roster spreadsheets into seed files for supabase/schema.sql:

    technicians_seed.csv / .sql                -> `technicians`
    technician_plant_access_seed.csv / .sql    -> `technician_plant_access`
    technician_certifications_seed.csv / .sql  -> `technician_certifications`

Two different source files, because each is the more authoritative shape for
what it produces:
  - Technician_Plant_Access.xlsx (horizontal AMP-number columns, one row per
    technician) drives `technicians` and `technician_plant_access`.
  - The original per-cert roster (Current_SPT_SMDT_list_CLEANED.xlsx, one row
    per technician per certification type) drives `technician_certifications`
    directly, rather than re-parsing the "Certification(s)" text blob the
    other file collapses those rows into.

Run supabase/schema.sql in the Supabase SQL Editor first to create the
tables. Then load the data either way:

  - Table Editor -> (table) -> Insert -> Import data from CSV, using the
    .csv files. This is the easiest path, but Studio's CSV importer has
    been unreliable on tables with composite primary keys / foreign keys
    (technician_plant_access and technician_certifications are both) -- if
    it rejects a file as "incompatible" with no useful detail, fall back to
    the matching .sql file instead of fighting the importer.
  - SQL Editor -> paste the whole .sql file -> Run. Slower to set up but
    goes through the same path that already worked for schema.sql. All
    three .sql files use ON CONFLICT DO NOTHING, so re-running them after
    the roster changes only inserts new rows -- it won't duplicate or error
    on ones already there. (It also won't *update* a changed row or drop
    a removed one; for that, clear the table first or write a real diff.)

These seed files contain real names -- do not commit them to git. Re-run
this script whenever either roster spreadsheet changes.

Usage:
    python3 build_technician_seed.py /path/to/Technician_Plant_Access.xlsx \\
        /path/to/Current_SPT_SMDT_list_CLEANED.xlsx [output_dir]
"""

import csv
import sys
from pathlib import Path

import openpyxl

ACCESS_SHEET_NAME = "Technician Plant Access"
ROSTER_SHEET_NAME = "Query1"

# Roster spreadsheet's exact certification strings -> the cert_type values
# technician_certifications.cert_type is constrained to (see schema.sql).
CERT_TYPE_MAP = {
    "Superpave Plant Technologist": "plant_tech",
    "Superpave Mix Design Technologist": "mix_design_tech",
}


def find_header_row(ws, label, max_row=10):
    """Anchor on a label cell rather than assuming a fixed row, in case a
    title or blank row gets added above the table later."""
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if cell.value == label:
                return cell.row
    raise ValueError(f"Could not find a {label!r} header cell in sheet {ws.title!r}")


def sql_string(value):
    """Postgres string literal, or NULL for an empty/missing value."""
    if value is None or value == "":
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def sql_date(value):
    if value is None:
        return "NULL"
    return "'" + value.strftime("%Y-%m-%d") + "'"


def read_technician_plant_access(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[ACCESS_SHEET_NAME]

    header_row = find_header_row(ws, "SM ID")
    headers = [c.value for c in ws[header_row]]
    col = {name: idx for idx, name in enumerate(headers) if name}
    amp_cols = [idx for name, idx in col.items() if isinstance(name, str) and name.startswith("AMP ")]

    technicians = []  # (sm_id, first, last, company, certs)
    access_rows = []  # (sm_id, amp_number)

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        sm_id = row[col["SM ID"]]
        if not sm_id:
            break  # first blank row ends the table; a notes block may follow it
        technicians.append(
            (sm_id, row[col["First Name"]], row[col["Last Name"]], row[col["Company"]], row[col["Certification(s)"]])
        )
        for idx in amp_cols:
            amp = row[idx]
            if amp:
                access_rows.append((sm_id, amp))

    return technicians, access_rows


def read_certifications(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[ROSTER_SHEET_NAME]

    header_row = find_header_row(ws, "SM ID")
    headers = [c.value for c in ws[header_row]]
    col = {name: idx for idx, name in enumerate(headers) if name}

    cert_rows = []  # (sm_id, cert_type, expires_on)
    unmapped = set()

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        sm_id = row[col["SM ID"]]
        if not sm_id:
            continue
        raw_cert = row[col["Certification Type"]]
        cert_type = CERT_TYPE_MAP.get(raw_cert)
        if cert_type is None:
            unmapped.add(raw_cert)
            continue
        cert_rows.append((sm_id, cert_type, row[col["Expiration Date"]]))

    if unmapped:
        raise ValueError(
            f"Unrecognized Certification Type value(s) in roster: {unmapped!r} - "
            f"add them to CERT_TYPE_MAP (and a matching value in schema.sql's "
            f"cert_type check constraint) before re-running."
        )

    return cert_rows


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    access_src = Path(sys.argv[1])
    roster_src = Path(sys.argv[2])
    out_dir = Path(sys.argv[3]) if len(sys.argv) > 3 else access_src.parent

    technicians, access_rows = read_technician_plant_access(access_src)
    cert_rows = read_certifications(roster_src)

    # --- CSV output (Table Editor import) ---
    technicians_csv = out_dir / "technicians_seed.csv"
    access_csv = out_dir / "technician_plant_access_seed.csv"
    certs_csv = out_dir / "technician_certifications_seed.csv"

    with open(technicians_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sm_id", "first_name", "last_name", "company", "certifications"])
        w.writerows(technicians)

    with open(access_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sm_id", "amp_number"])
        w.writerows(access_rows)

    with open(certs_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sm_id", "cert_type", "expires_on"])
        w.writerows((sm_id, cert_type, exp.strftime("%Y-%m-%d") if exp else "") for sm_id, cert_type, exp in cert_rows)

    # --- SQL output (SQL Editor fallback) ---
    technicians_sql = out_dir / "technicians_seed.sql"
    access_sql = out_dir / "technician_plant_access_seed.sql"
    certs_sql = out_dir / "technician_certifications_seed.sql"

    with open(technicians_sql, "w", encoding="utf-8") as f:
        f.write("insert into technicians (sm_id, first_name, last_name, company, certifications)\nvalues\n")
        f.write(
            ",\n".join(
                f"  ({sql_string(sm_id)}, {sql_string(first)}, {sql_string(last)}, {sql_string(company)}, {sql_string(certs)})"
                for sm_id, first, last, company, certs in technicians
            )
        )
        f.write("\non conflict (sm_id) do nothing;\n")
        # Central Office Materials reviewers see every plant (see
        # supabase/effective_plant_access.sql). Set the flag by company so a
        # re-seed that adds a new Central Office person grants it too,
        # instead of someone remembering to flip it by hand.
        f.write(
            "\nupdate technicians set all_plants = true\n"
            " where company = 'Central Office Materials' and not all_plants;\n"
        )

    with open(access_sql, "w", encoding="utf-8") as f:
        f.write("insert into technician_plant_access (sm_id, amp_number)\nvalues\n")
        f.write(",\n".join(f"  ({sql_string(sm_id)}, {sql_string(amp)})" for sm_id, amp in access_rows))
        f.write("\non conflict (sm_id, amp_number) do nothing;\n")

    with open(certs_sql, "w", encoding="utf-8") as f:
        f.write("insert into technician_certifications (sm_id, cert_type, expires_on)\nvalues\n")
        f.write(
            ",\n".join(
                f"  ({sql_string(sm_id)}, {sql_string(cert_type)}, {sql_date(exp)})"
                for sm_id, cert_type, exp in cert_rows
            )
        )
        f.write("\non conflict (sm_id, cert_type) do update set expires_on = excluded.expires_on;\n")

    print(f"Wrote {technicians_csv} and {technicians_sql} ({len(technicians)} technicians)")
    print(f"Wrote {access_csv} and {access_sql} ({len(access_rows)} technician-plant access rows)")
    print(f"Wrote {certs_csv} and {certs_sql} ({len(cert_rows)} certification rows)")


if __name__ == "__main__":
    main()
